import re
import os
import asyncio
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import Application, CommandHandler, MessageHandler, filters, CallbackQueryHandler, ContextTypes, ConversationHandler
from openpyxl import Workbook, load_workbook
from telegram.error import TelegramError

TOKEN = "8137824808:AAG3Kua_oKTigk3Dt4LEvtbhf76OzCial3A"
SITE_URL = "https://tkaniruna.ru/"
VIDEO_PATH = "src/guide.mp4" 
PDF_PATH = "src/Как пошить платье.pdf"     
EXCEL_FILE = "users.xlsx"
ADMIN_USER_IDS = [1985211012]  # id администратора

WELCOME_TEXT = (
    "Привет, {username}!\n\n"
    "Добро пожаловать в бот RUNATEX!\n\n"
    "Здесь мы делимся:\n"
    "— полезными советами по выбору тканей и уходу за ними\n"
    "— вдохновением для ваших творческих проектов\n"
    "— новостями о самых свежих поступлениях и акциях\n\n"
    "Оставайтесь на связи — будет красиво, полезно и по делу."
)

ASK_EMAIL_TEXT = "Пожалуйста, укажи свою почту для регистрации:"

BROADCAST_TEXT, BROADCAST_PHOTO, BROADCAST_LINK = range(3)

def is_valid_email(email):
    return re.match(r"[^@]+@[^@]+\.[^@]+", email)

def save_user_to_excel(username, email, user_id):
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(["Username", "Email", "UserID"])
        wb.save(EXCEL_FILE)
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([username, email, user_id])
    wb.save(EXCEL_FILE)

def is_user_registered(username):
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == username:
            return True
    return False

def get_all_user_ids():
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    user_ids = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if len(row) > 2 and row[2]:
            user_ids.add(row[2])
    return list(user_ids)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    username = update.effective_user.username or update.effective_user.first_name or "unknown"
    if is_user_registered(username):
        await update.message.reply_text(
            "Вы уже зарегистрированы! Выберите действие:",
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("Перейти на сайт", url=SITE_URL)],
                [InlineKeyboardButton("Получить обучающее видео", callback_data="video")],
                [InlineKeyboardButton("Наши соцсети", url="https://taplink.cc/tkani_runa")]
            ])
        )
        return
    username_for_greeting = update.effective_user.first_name or "друг"
    await update.message.reply_text(
        WELCOME_TEXT.format(username=username_for_greeting),
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("Зарегистрироваться", callback_data="register")]
        ])
    )

async def button_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    if query.data == "register":
        await query.message.reply_text(ASK_EMAIL_TEXT)
        context.user_data["awaiting_email"] = True
    elif query.data == "site":
        await query.message.reply_text(f"Переходи на сайт: {SITE_URL}")
    elif query.data == "video":
        await query.message.reply_video(open(VIDEO_PATH, "rb"))
        await query.message.reply_document(open(PDF_PATH, "rb"))

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if context.user_data.get("awaiting_email"):
        email = update.message.text
        username = update.effective_user.username or update.effective_user.first_name or "unknown"
        user_id = update.effective_user.id
        if is_user_registered(username):
            context.user_data["awaiting_email"] = False
            await update.message.reply_text(
                "Вы уже зарегистрированы! Если хотите изменить почту, обратитесь к администратору."
            )
            return
        if is_valid_email(email):
            context.user_data["awaiting_email"] = False
            save_user_to_excel(username, email, user_id)
            await update.message.reply_text(
                "Спасибо за регистрацию! Выберите действие:",
                reply_markup=InlineKeyboardMarkup([
                    [InlineKeyboardButton("Перейти на сайт", url=SITE_URL)],
                    [InlineKeyboardButton("Получить обучающее видео", callback_data="video")],
                    [InlineKeyboardButton("Наши соцсети", url="https://taplink.cc/tkani_runa")]
                ])
            )
        else:
            await update.message.reply_text("Почта некорректна, попробуйте еще раз:")

async def broadcast_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id not in ADMIN_USER_IDS:
        await update.message.reply_text("У вас нет доступа к этой команде.")
        return ConversationHandler.END
    await update.message.reply_text("Введите текст рассылки:")
    return BROADCAST_TEXT

async def broadcast_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['broadcast_text'] = update.message.text
    await update.message.reply_text("Хотите добавить фото? Пришлите фото или напишите 'нет'.")
    return BROADCAST_PHOTO

async def broadcast_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.photo:
        photo_file_id = update.message.photo[-1].file_id
        context.user_data['broadcast_photo'] = photo_file_id
        await update.message.reply_text("Хотите добавить кнопку-ссылку? Пришлите ссылку или напишите 'нет'.")
        return BROADCAST_LINK
    elif update.message.text and update.message.text.lower() == 'нет':
        context.user_data['broadcast_photo'] = None
        await update.message.reply_text("Хотите добавить кнопку-ссылку? Пришлите ссылку или напишите 'нет'.")
        return BROADCAST_LINK
    else:
        await update.message.reply_text("Пожалуйста, пришлите фото или напишите 'нет'.")
        return BROADCAST_PHOTO

async def broadcast_link(update: Update, context: ContextTypes.DEFAULT_TYPE):
    link = update.message.text
    if link.lower() == 'нет':
        link = None
    context.user_data['broadcast_link'] = link
    user_ids = get_all_user_ids()
    count = 0
    for uid in user_ids:
        try:
            if context.user_data['broadcast_photo']:
                reply_markup = None
                if link:
                    reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton("Перейти", url=link)]])
                await context.bot.send_photo(
                    chat_id=uid,
                    photo=context.user_data['broadcast_photo'],
                    caption=context.user_data['broadcast_text'],
                    reply_markup=reply_markup
                )
            else:
                reply_markup = None
                if link:
                    reply_markup = InlineKeyboardMarkup([[InlineKeyboardButton("Перейти", url=link)]])
                await context.bot.send_message(
                    chat_id=uid,
                    text=context.user_data['broadcast_text'],
                    reply_markup=reply_markup
                )
            count += 1
        except Exception:
            continue
    await update.message.reply_text(f"Рассылка завершена. Сообщение отправлено {count} пользователям.")
    return ConversationHandler.END

async def broadcast_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Рассылка отменена.")
    return ConversationHandler.END

async def get_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    if user_id in ADMIN_USER_IDS:
        if os.path.exists(EXCEL_FILE):
            await update.message.reply_document(open(EXCEL_FILE, "rb"))
        else:
            await update.message.reply_text("Файл с пользователями пока не создан.")
    else:
        await update.message.reply_text("У вас нет доступа к этой команде.")

async def main():
    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("getusers", get_users))
    app.add_handler(ConversationHandler(
        entry_points=[CommandHandler("broadcast", broadcast_start)],
        states={
            BROADCAST_TEXT: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_text)],
            BROADCAST_PHOTO: [MessageHandler((filters.PHOTO | filters.TEXT) & ~filters.COMMAND, broadcast_photo)],
            BROADCAST_LINK: [MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_link)],
        },
        fallbacks=[CommandHandler("cancel", broadcast_cancel)],
    ))
    app.add_handler(CallbackQueryHandler(button_handler))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))

    await app.initialize()
    await app.start()
    await app.updater.start_polling()
    await app.updater.idle()

if __name__ == "__main__":
    asyncio.run(main())